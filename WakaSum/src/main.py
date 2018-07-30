import openpyxl as xl
from openpyxl import load_workbook
from extract import all_data

def main():

    path = input("Select the path with the table and the text files: ")
    load(path)
    try:
        write(path)
        print("Done!")
    except PermissionError:
        print("This file is used by another program.")
    

def load(path):
    """Loading the WorkBook and extracting the data from the text files"""
    global wb, sheet, data

    wb = load_workbook(path + "/all_data.xlsx")
    sheet = wb["Sheet1"]
    data = all_data(path)

def write(path):
    """Writing to the Excel table"""

    sheet["B2"].value, sheet["C2"].value = min_to_hour(data["Total"])
    r = 2
    c = 1
    for el in data:
        if el != "Total":
            r = 2
            c += 3
            for i in data[el]:
                sheet.cell(row=r, column= c).value = i
                sheet.cell(row=r, column= c + 1).value, sheet.cell(row=r, column= c + 2).value = min_to_hour(data[el][i])
                r += 1

    wb.save(path + "/all_data.xlsx")

def min_to_hour(mins):
    """Format mins to hours and mins"""
    return (int((mins - mins % 60) / 60), mins % 60)

if __name__ == "__main__":
    main()